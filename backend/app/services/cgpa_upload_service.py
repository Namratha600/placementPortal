"""
Bulk CGPA update from an uploaded CSV/Excel file.

File columns (header row required): Register Number, CGPA
(various common spellings accepted). Each row is matched to a student by
register number and their CGPA is updated. Valid rows are applied in a single
transaction; unmatched / invalid rows are skipped and reported.
"""
import csv
import io
from typing import List, Dict, Optional, Tuple

from fastapi import HTTPException
from sqlalchemy import func
from sqlalchemy.orm import Session

from app.models.student import Student

HEADER_ALIASES = {
    "register_number": {"register number", "registernumber", "reg no", "regno",
                        "roll number", "rollnumber", "roll no", "rollno", "register_no"},
    "cgpa": {"cgpa", "gpa", "cgpa (10)", "cgpa10"},
}
REQUIRED = {"register_number", "cgpa"}
ALLOWED_EXTENSIONS = {".csv", ".xlsx"}


def _normalize_header(raw: str) -> Optional[str]:
    key = (raw or "").strip().lower()
    for field, aliases in HEADER_ALIASES.items():
        if key == field or key in aliases:
            return field
    return None


def parse_file(filename: str, content: bytes) -> List[Dict[str, str]]:
    lower = (filename or "").lower()
    ext = lower[lower.rfind("."):] if "." in lower else ""
    if ext not in ALLOWED_EXTENSIONS:
        raise HTTPException(status_code=400,
                            detail=f"Unsupported file type '{ext}'. Upload a .csv or .xlsx file.")

    all_rows = _parse_csv(content) if ext == ".csv" else _parse_xlsx(content)
    dict_rows = _rows_to_dicts(all_rows)
    if not dict_rows:
        raise HTTPException(status_code=400, detail="The file has no data rows.")
    return dict_rows


def _parse_csv(content: bytes) -> List[List[str]]:
    text = content.decode("utf-8-sig", errors="replace")
    return [r for r in csv.reader(io.StringIO(text))]


def _parse_xlsx(content: bytes) -> List[List[str]]:
    try:
        from openpyxl import load_workbook
    except ImportError:
        raise HTTPException(status_code=500, detail="openpyxl is not installed.")
    wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    ws = wb.active
    rows = [["" if c is None else str(c) for c in row]
            for row in ws.iter_rows(values_only=True)]
    wb.close()
    return rows


def _rows_to_dicts(all_rows: List[List[str]]) -> List[Dict[str, str]]:
    header_idx = None
    header_map = []
    for i, row in enumerate(all_rows):
        mapped = [_normalize_header(str(c)) for c in row]
        if all(req in mapped for req in REQUIRED):
            header_idx, header_map = i, mapped
            break
    if header_idx is None:
        raise HTTPException(
            status_code=400,
            detail="File must contain a Register Number column and a CGPA column.",
        )

    out = []
    for row in all_rows[header_idx + 1:]:
        if all(str(c).strip() == "" for c in row):
            continue
        rec = {}
        for col_idx, field in enumerate(header_map):
            if field and col_idx < len(row):
                rec[field] = str(row[col_idx]).strip()
        out.append(rec)
    return out


def _validate(raw: Dict[str, str]) -> Tuple[Optional[dict], Optional[str]]:
    reg = (raw.get("register_number") or "").strip()
    cgpa_raw = (raw.get("cgpa") or "").strip()
    if not reg:
        return None, "Missing register number"
    if not cgpa_raw:
        return None, "Missing CGPA"
    try:
        cgpa = float(cgpa_raw)
    except ValueError:
        return None, f"Invalid CGPA '{cgpa_raw}'"
    if cgpa < 0 or cgpa > 10:
        return None, f"CGPA out of range (0-10): {cgpa}"
    return {"register_number": reg, "cgpa": round(cgpa, 2)}, None


def bulk_update_cgpa(db: Session, rows: List[Dict[str, str]]) -> dict:
    updated = skipped = 0
    errors = []

    try:
        for i, raw in enumerate(rows, start=1):
            clean, reason = _validate(raw)
            if reason:
                skipped += 1
                errors.append({"row": i, "reason": reason})
                continue

            student = (
                db.query(Student)
                .filter(func.lower(Student.register_number) == clean["register_number"].lower())
                .first()
            )
            if not student:
                skipped += 1
                errors.append({"row": i,
                               "reason": f"No student with register number {clean['register_number']}"})
                continue

            student.cgpa = clean["cgpa"]
            updated += 1

        db.commit()
    except Exception as e:
        db.rollback()
        raise HTTPException(status_code=500,
                            detail=f"Update failed and was rolled back. No changes saved. ({e})")

    return {"updated": updated, "skipped": skipped,
            "total_rows": len(rows), "errors": errors}