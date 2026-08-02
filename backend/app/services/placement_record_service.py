"""
Business logic for placement records:
  - parse an uploaded CSV or XLSX file into normalized rows
  - validate each row (collecting per-row errors, never crashing on one bad row)
  - resolve/create the Company for each row (by name)
  - resolve the student_id by matching roll_number -> students.register_number
  - UPSERT by (roll_number, company_id) inside a single transaction
  - filtered + paginated retrieval for the admin 'Get Records' view

Only stdlib `csv` + `openpyxl` are used for parsing (no pandas) to stay safe
on Python 3.14 / Windows wheels.
"""
import csv
import io
from datetime import date
from typing import List, Dict, Tuple, Optional

from fastapi import HTTPException
from sqlalchemy import func
from sqlalchemy.orm import Session

from app.models.company import Company
from app.models.placement_record import PlacementRecord


# ---------------------------------------------------------------------------
# Header handling
# ---------------------------------------------------------------------------
# Map many possible spellings in the uploaded file to our internal field names.
HEADER_ALIASES = {
    "roll_number": {"roll number", "rollnumber", "roll no", "rollno",
                    "register number", "registernumber", "reg no", "student id", "studentid"},
    "student_name": {"student name", "name", "studentname"},
    "branch": {"branch", "department", "dept"},
    "graduation_year": {"graduation year", "grad year", "passing year",
                        "year", "graduationyear"},
    "company": {"company", "company name", "companyname"},
    "role": {"role", "designation", "position", "job role"},
    "package": {"package", "ctc", "package lpa", "salary", "package (lpa)"},
    "placement_date": {"placement date", "date", "offer date", "placementdate"},
}

REQUIRED_FIELDS = {"roll_number", "company"}   # minimum needed to make a record
ALLOWED_EXTENSIONS = {".csv", ".xlsx"}


def _normalize_header(raw: str) -> Optional[str]:
    key = (raw or "").strip().lower()
    for field, aliases in HEADER_ALIASES.items():
        if key == field or key in aliases:
            return field
    return None


# ---------------------------------------------------------------------------
# File parsing -> list of raw string dicts keyed by our internal field names
# ---------------------------------------------------------------------------
def parse_file(filename: str, content: bytes) -> List[Dict[str, str]]:
    lower = (filename or "").lower()
    ext = lower[lower.rfind("."):] if "." in lower else ""

    if ext not in ALLOWED_EXTENSIONS:
        raise HTTPException(
            status_code=400,
            detail=f"Unsupported file type '{ext}'. Upload a .csv or .xlsx file.",
        )

    if ext == ".csv":
        rows = _parse_csv(content)
    else:
        rows = _parse_xlsx(content)

    if not rows:
        raise HTTPException(status_code=400, detail="The file has no data rows.")
    return rows


def _parse_csv(content: bytes) -> List[Dict[str, str]]:
    text = content.decode("utf-8-sig", errors="replace")  # utf-8-sig strips Excel BOM
    reader = csv.reader(io.StringIO(text))
    all_rows = [r for r in reader]
    return _rows_to_dicts(all_rows)


def _parse_xlsx(content: bytes) -> List[Dict[str, str]]:
    try:
        from openpyxl import load_workbook
    except ImportError:
        raise HTTPException(
            status_code=500,
            detail="openpyxl is not installed. Run: pip install openpyxl",
        )
    wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    ws = wb.active
    all_rows = []
    for row in ws.iter_rows(values_only=True):
        all_rows.append(["" if c is None else str(c) for c in row])
    wb.close()
    return _rows_to_dicts(all_rows)


def _rows_to_dicts(all_rows: List[List[str]]) -> List[Dict[str, str]]:
    """First non-empty row = header. Remaining rows -> dicts keyed by field name."""
    # find header row (first row that has at least one recognizable header)
    header_idx = None
    for i, row in enumerate(all_rows):
        mapped = [_normalize_header(str(c)) for c in row]
        if any(m in REQUIRED_FIELDS for m in mapped):
            header_idx = i
            header_map = mapped
            break
    if header_idx is None:
        raise HTTPException(
            status_code=400,
            detail="Could not find required columns. The file must contain at "
                   "least a Roll Number column and a Company column.",
        )

    dict_rows: List[Dict[str, str]] = []
    for row in all_rows[header_idx + 1:]:
        if all((str(c).strip() == "" for c in row)):
            continue  # skip fully blank lines
        record: Dict[str, str] = {}
        for col_idx, field in enumerate(header_map):
            if field and col_idx < len(row):
                record[field] = str(row[col_idx]).strip()
        dict_rows.append(record)
    return dict_rows


# ---------------------------------------------------------------------------
# Row validation
# ---------------------------------------------------------------------------
def _validate_and_coerce(raw: Dict[str, str]) -> Tuple[Optional[dict], Optional[str]]:
    """Return (clean_dict, None) on success or (None, reason) on failure."""
    roll = (raw.get("roll_number") or "").strip()
    company_name = (raw.get("company") or "").strip()

    if not roll:
        return None, "Missing roll number"
    if not company_name:
        return None, "Missing company name"

    clean = {
        "roll_number": roll,
        "company_name": company_name,
        "student_name": (raw.get("student_name") or "").strip() or None,
        "branch": (raw.get("branch") or "").strip() or None,
        "role": (raw.get("role") or "").strip() or None,
        "graduation_year": None,
        "package": None,
        "placement_date": None,
    }

    gy = (raw.get("graduation_year") or "").strip()
    if gy:
        try:
            clean["graduation_year"] = int(float(gy))
        except ValueError:
            return None, f"Invalid graduation year '{gy}'"

    pkg = (raw.get("package") or "").strip().replace(",", "")
    if pkg:
        try:
            clean["package"] = float(pkg)
        except ValueError:
            return None, f"Invalid package '{pkg}'"

    pd_raw = (raw.get("placement_date") or "").strip()
    if pd_raw:
        parsed = _parse_date(pd_raw)
        if parsed is None:
            return None, f"Invalid placement date '{pd_raw}'"
        clean["placement_date"] = parsed

    return clean, None


def _parse_date(value: str) -> Optional[date]:
    from datetime import datetime
    for fmt in ("%Y-%m-%d", "%d-%m-%Y", "%d/%m/%Y", "%m/%d/%Y"):
        try:
            return datetime.strptime(value, fmt).date()
        except ValueError:
            continue
    return None


# ---------------------------------------------------------------------------
# Merge (transactional upsert)
# ---------------------------------------------------------------------------
def merge_records(db: Session, rows: List[Dict[str, str]]) -> dict:
    inserted = updated = skipped = 0
    errors = []
    companies_created = []

    # caches to avoid repeat lookups within one upload
    company_cache: Dict[str, Company] = {}
    student_cache: Dict[str, Optional[int]] = {}

    try:
        for i, raw in enumerate(rows, start=1):
            clean, reason = _validate_and_coerce(raw)
            if reason:
                skipped += 1
                errors.append({"row": i, "reason": reason})
                continue

            # resolve (or create) company by name
            cname = clean["company_name"]
            company = company_cache.get(cname.lower())
            if company is None:
                company = (
                    db.query(Company)
                    .filter(func.lower(Company.name) == cname.lower())
                    .first()
                )
                if company is None:
                    company = Company(name=cname)
                    db.add(company)
                    db.flush()          # get company.id without committing yet
                    companies_created.append(cname)
                company_cache[cname.lower()] = company

            # resolve student_id by roll_number -> students.register_number
            roll = clean["roll_number"]
            if roll not in student_cache:
                student_cache[roll] = _lookup_student_id(db, roll)
            student_id = student_cache[roll]

            # upsert by (roll_number, company_id)
            existing = (
                db.query(PlacementRecord)
                .filter(
                    PlacementRecord.roll_number == roll,
                    PlacementRecord.company_id == company.id,
                )
                .first()
            )
            if existing:
                existing.student_id = student_id
                existing.student_name = clean["student_name"]
                existing.branch = clean["branch"]
                existing.graduation_year = clean["graduation_year"]
                existing.role = clean["role"]
                existing.package = clean["package"]
                existing.placement_date = clean["placement_date"]
                updated += 1
            else:
                db.add(PlacementRecord(
                    company_id=company.id,
                    student_id=student_id,
                    roll_number=roll,
                    student_name=clean["student_name"],
                    branch=clean["branch"],
                    graduation_year=clean["graduation_year"],
                    role=clean["role"],
                    package=clean["package"],
                    placement_date=clean["placement_date"],
                ))
                inserted += 1

        db.commit()   # single transaction for the whole valid batch
    except Exception as e:
        db.rollback()
        raise HTTPException(
            status_code=500,
            detail=f"Upload failed and was rolled back. No records were saved. ({e})",
        )

    return {
        "inserted": inserted,
        "updated": updated,
        "skipped": skipped,
        "total_rows": len(rows),
        "companies_created": companies_created,
        "errors": errors,
    }


def _lookup_student_id(db: Session, roll_number: str) -> Optional[int]:
    """Best-effort link to an existing student by register_number."""
    from app.models.student import Student
    student = (
        db.query(Student)
        .filter(func.lower(Student.register_number) == roll_number.lower())
        .first()
    )
    return student.id if student else None


# ---------------------------------------------------------------------------
# Retrieval (admin 'Get Records' with filters + pagination)
# ---------------------------------------------------------------------------
def get_records(
    db: Session,
    year: Optional[int] = None,
    branch: Optional[str] = None,
    company_id: Optional[int] = None,
    page: int = 1,
    page_size: int = 25,
) -> dict:
    page = max(page, 1)
    page_size = min(max(page_size, 1), 200)

    q = db.query(PlacementRecord)
    if year is not None:
        q = q.filter(PlacementRecord.graduation_year == year)
    if branch:
        q = q.filter(func.lower(PlacementRecord.branch) == branch.lower())
    if company_id is not None:
        q = q.filter(PlacementRecord.company_id == company_id)

    total = q.count()
    items = (
        q.order_by(PlacementRecord.graduation_year.desc(),
                   PlacementRecord.id.desc())
        .offset((page - 1) * page_size)
        .limit(page_size)
        .all()
    )
    return {"total": total, "page": page, "page_size": page_size, "items": items}

def build_records_xlsx(db, year=None, branch=None, company_id=None) -> bytes:
    """Build an .xlsx of all placement records matching the filters (no pagination)."""
    from openpyxl import Workbook
    from io import BytesIO

    q = db.query(PlacementRecord)
    if year is not None:
        q = q.filter(PlacementRecord.graduation_year == year)
    if branch:
        q = q.filter(func.lower(PlacementRecord.branch) == branch.lower())
    if company_id is not None:
        q = q.filter(PlacementRecord.company_id == company_id)
    records = q.order_by(
        PlacementRecord.graduation_year.desc(), PlacementRecord.id.desc()
    ).all()

    cids = {r.company_id for r in records}
    cmap = {}
    if cids:
        for cid, cname in db.query(Company.id, Company.name).filter(Company.id.in_(cids)).all():
            cmap[cid] = cname

    wb = Workbook()
    ws = wb.active
    ws.title = "Placement Records"
    ws.append(["Roll Number", "Student Name", "Branch", "Graduation Year",
               "Company", "Role", "Package", "Placement Date"])
    for r in records:
        ws.append([
            r.roll_number, r.student_name, r.branch, r.graduation_year,
            cmap.get(r.company_id, r.company_id), r.role,
            float(r.package) if r.package is not None else None,
            r.placement_date.isoformat() if r.placement_date else None,
        ])

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()    